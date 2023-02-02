#! usr/bin/python3

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, \
    PatternFill, Border, Side, Color

# import tkinter as tk
# from tkinter import ttk
# from tkinter import filedialog

import os
import logging
import datetime
import argparse
import subprocess
# import shelve

logging.basicConfig(level=logging.DEBUG, format='%(levelname)s - %(message)s')
# logging.disable()

parser = argparse.ArgumentParser()
parser.add_argument('-p', '--product', default='GASOIL',
                    help='Indiquer le type du product')
parser.add_argument('-a', '--adjust', default=0, help='Previous solde')
args = vars(parser.parse_args())

filepath = '/home/chrisbal/Downloads/Carburant 2022.xlsx'

# filepath = shelve
# root = tk.Tk()
# filepath = filedialog.askopenfilename()
# root.mainloop()
os.chdir(os.path.dirname(filepath))
os.chdir(os.path.dirname(filepath))
wb = load_workbook(filepath)

gasoil_ws = wb['GASOIL']
essence_ws = wb['ESSENCE']
wb1 = Workbook()
wb2 = Workbook()
m = 1
n = 3


def fiche_stock(ws):
    global m
    data = {}
    m = 3
    fiche_ws = wb1[wb1.sheetnames[0]]
    fiche_ws.title = 'Fiche'
    fiche_ws['A1'] = 'FICHE DE SUIVI DU ' + args['product'].upper()
    fiche_ws.merge_cells('A1:H1')
    a = fiche_ws['A1']

    headers = ['date', 'stock', 'stock in',
               'total', 'stock out', 'balance', 'Signature', 'Observation']
    for j in range(len(headers)):
        cell = fiche_ws.cell(row=m, column=j + 1, value=headers[j].upper())
        cell.font = Font(bold=True,
                         color=Color(indexed=63),
                         size=12)
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   shrink_to_fit=True)
        cell.border = Border(top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             left=Side(border_style='thin'))
        dims = {}
        if cell.value and cell.value != 'A':
            dims.setdefault(cell.column_letter, len(str(cell.value)))

        for letter, width1 in dims.items():
            fiche_ws.column_dimensions[letter].width = width1 + 5

    def convert(value):
        try:
            if value is None:
                value = 0
            else:
                value = float(value)
        except TypeError:
            pass
        except ValueError:
            value = 0
        return value

    for row in ws.iter_rows(min_row=4):
        date = row[0].value
        entr = row[5].value
        out = row[6].value
        date1 = row[8].value
        out1 = row[12].value

        data.setdefault(date, {})
        data.setdefault(date1, {})
        data[date1].setdefault('out', 0)
        data[date1].setdefault('inp', 0)
        data[date].setdefault('inp', 0)
        data[date].setdefault('out', 0)
        data[date]['inp'] += convert(entr)
        data[date]['out'] += convert(out)
        data[date1]['out'] += convert(out1)

    deletable = []
    for_analyze = {}
    data_items = data.items()
    for date, qte in data_items:
        if date is None:
            if not(bool(qte.get('inp', None)) or bool(qte.get('out', None))):
                deletable.append(date)
        if not isinstance(date, datetime.datetime):
            for_analyze.setdefault(date, data[date])
            if date not in deletable:
                deletable.append(date)

    for delete in deletable:
        del data[delete]

    datas = sorted(data.items())

    for date, qte in datas:
        fiche_ws[f'A{m + 1}'].value = date.strftime('%d/%m/%Y')
        fiche_ws[f'C{m + 1}'].value = qte.get('inp', 0)
        fiche_ws[f'E{m + 1}'].value = qte.get('out', 0)
        m += 1

    for row in fiche_ws.iter_rows(min_row=4):
        x = row[0].row
        fiche_ws.row_dimensions[x].height = 20

        if int(row[2].value) > 0:
            for cell in row:
                cell.fill = PatternFill(fill_type='solid',
                                        end_color=Color(indexed=0),
                                        start_color=Color(indexed=31))

        for cell in row:
            d = '00111111'
            cell.font = Font(name='Verdana', color=Color(indexed=63))
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
            cell.border = Border(left=Side(border_style='dotted',
                                           color=d),
                                 right=Side(border_style='dotted',
                                            color=d),
                                 bottom=Side(border_style='dotted',
                                             color=d))

        fiche_ws[f'B{x}'] = args['adjust'] if x == 4 else f'=F{x - 1}'
        fiche_ws[f'D{x}'] = f'=B{x}+C{x}'
        fiche_ws[f'F{x}'] = f'=D{x}-E{x}'

    a.font = Font(name='Verdana',
                  size=18,
                  bold=True,
                  color='003366FF')
    a.alignment = Alignment(horizontal='center',
                            vertical='center',
                            shrink_to_fit=True)
    fiche_ws.column_dimensions['A'].width = 16
    a.border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      top=Side(border_style='medium'),
                      bottom=Side(border_style='medium'))
    fiche_ws.row_dimensions[1].height = 30
    wb1.save('fiche.xlsx')
    subprocess.Popen(['open', 'fiche.xlsx'])


def monthly_rapport(ws):
    global n
    months = ['janvier',
              'fevrier',
              'mars',
              'avril',
              'mai',
              'juin',
              'juillet',
              'aout',
              'septembre',
              'octobre',
              'novembre',
              'decembre']
    headers1 = ['Date', 'Temoin', 'Motif', 'Autorisation', 'Client', 'Qte']
    headers2 = ['Date', 'Temoin', 'Autorisation', 'Qte', 'km A', 'km B', 'Diff km', 'Conso. moyenne']
    data = {'a': {}, 'b': {}}
    big_cons = {}
    inp1 = ['date', 'name', 'motive', 'authorization', 'client', 'in', 'out']
    inp2 = ['date', 'name', 'engin', 'auth', 'out', 'kmA', 'kmB', 'diff', 'cons']
    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G',
                'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
                'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    sheets = {}
    prev_solde = {}
    sub_total = {}
    supply = {}
    for j in range(len(months)):
        sheets.setdefault(j + 1, months[j])
        sub_total.setdefault(j + 1, 0)
        supply.setdefault(j + 1, 0)
        prev_solde.setdefault(j + 1, None)
        if j == 0:
            ws1 = wb2[wb2.sheetnames[0]]
            ws1.title = months[0]
            continue
        wb2.create_sheet(months[j])

    for key in sheets:
        ws_active = wb2[sheets[key]]
        a = ws_active['A1']
        a.value = 'Consommations'

        j = 0
        for col in ws_active.iter_cols(max_col=len(headers1)):
            col_letter = col[0].column_letter
            ws_active[f'{col_letter}2'] = headers1[j]
            if j == len(headers1) - 1:
                ws_active[f'{col_letter}3'].value = 0
                sub_total[key] = ws_active[f'{col_letter}3']
            j += 1

    def write_cons_data(data_obj):
        sum_min_row = 0
        all_data = {}
        for engin in data_obj:
            for date in data_obj[engin]:
                try:
                    worksheet = wb2[sheets[date.month]]
                    all_data.setdefault(date.month, dict())
                    all_data[date.month].setdefault('engin', set())
                    all_data[date.month].setdefault('first_row', 0)
                    all_data[date.month].setdefault('engin_cons', dict())
                    all_data[date.month]['engin_cons'].setdefault(engin, set())
                    row_num = worksheet.max_row + 2
                    if engin not in all_data[date.month].get('engin', set()):
                        worksheet.insert_rows(idx=row_num)
                        title = worksheet[f'A{row_num}']
                        title.value = engin.upper()
                        title.font = Font(color='004C6ef5', size=18)
                        title.fill = PatternFill(fill_type='solid', start_color='00DEE2E6')
                        worksheet.merge_cells(f'A{row_num}:{alphabet[len(inp2)]}{row_num}')
                        if len(all_data[date.month].get('engin', set())) == 0:
                            all_data[date.month]['first_row'] = row_num
                        for i in range(len(headers2)):
                            letter = alphabet[i]
                            header = worksheet[f'{letter}{row_num + 1}']
                            header.value = headers2[i]
                            header.font = Font(size=12, bold=True, color='00343A40')
                            header.fill = PatternFill(fill_type='solid', start_color='00BAC8FF')
                            if i == 2:
                                worksheet[f'{letter}{row_num + 2}'].value = 'total'.upper()
                            if i >= 3:
                                worksheet[f'{letter}{row_num + 2}'].value = 0
                            if i == 3:
                                all_data[date.month]['engin_cons'][engin].add(worksheet[f'{letter}{row_num + 2}'])
                        sum_min_row = worksheet.max_row
                        all_data[date.month]['engin'].add(engin)

                    for name in data_obj[engin][date]:
                        related_data = [date, name]
                        for rel in data_obj[engin][date][name]:
                            related_data.append(data_obj[engin][date][name][rel])
                        while len(related_data) < len(headers2):
                            related_data.append(None)
                        worksheet.insert_rows(idx=worksheet.max_row)
                        row_num_wr = worksheet.max_row - 1
                        for i in range(len(headers2)):
                            letter = alphabet[i]
                            if i == 0:
                                worksheet[f'{letter}{row_num_wr}'].value = related_data[i].strftime('%d/%m/%Y')
                                continue
                            worksheet[f'{letter}{row_num_wr}'].value = related_data[i]
                        sum_max_row = worksheet.max_row

                        for sum_cell in all_data[date.month]['engin_cons'][engin]:
                            cell_col = sum_cell.column_letter
                            sum_cell.value = f'=SUM({cell_col}{sum_min_row}:{cell_col}{sum_max_row - 1})'
                except AttributeError:
                    pass
        for mon in all_data:
            worksheet = wb2[sheets[mon]]
            move_row = worksheet.max_row
            diff = move_row - all_data[mon].get('first_row', 0)
            worksheet.move_range(f"A{all_data[mon]['first_row']}:{alphabet[len(inp2)]}{move_row}",
                                 rows=-(all_data[mon].get('first_row', 1) - 1),
                                 cols=len(inp1), translate=True)
            total = '=SUM('

            synth_row = diff + 2

            synth_headers = {'Synthese': None,
                             'Designation': 'Quantite'}
            tot = {}
            i = 0
            for equip, cons in all_data[mon]['engin_cons'].items():
                tot.setdefault(equip, list(cons)[0].value)
                total += f'J{synth_row + i + 3}:'
                i += 1
            total = total[:len(total)-1] + ')'
            cons_total = f'{total} + {str(sub_total[mon].value)[1:]}' \
                if sub_total[mon].value != 0 else f'{total}'
            other_conso = sub_total[mon].value[1:] if sub_total[mon].value != 0 else '0'
            synth_rest = {'Autres': sub_total[mon].value,
                          'Total': cons_total,
                          'Entrees': supply[mon],
                          'Solde precedent': prev_solde[mon],
                          'Solde': f'={supply[mon]}+{prev_solde[mon][1:]}-{total[1:]}-{other_conso}'
                          if prev_solde[mon] is not None else f'={supply[mon]}-{total[1:]}-{other_conso}'
                          }

            synthese = synth_headers | tot | synth_rest
            logging.debug(prev_solde[mon])
            i = 0
            solde = None
            for designation, synth_value in synthese.items():
                designation_cell = worksheet[f'I{synth_row + 1 + i}']
                value_cell = worksheet[f'J{synth_row + 1 + i}']
                designation_cell.value = designation
                value_cell.value = synth_value
                if designation == 'Solde':
                    solde = value_cell
                i += 1

            try:
                solde_grid = f'{solde.column_letter}{solde.row}'
                prev_solde[mon + 1] = f'={sheets[mon]}!{solde_grid}'
            except KeyError:
                pass

            for _row in worksheet.iter_rows():
                for _cell in _row:
                    # _cell.font = Font(name='Verdana', size=11, color=Color(indexed=63))
                    _cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            head1 = worksheet['A1']
            head1.font = Font(color='003366FF', name='Verdana',
                              size=18, bold=True)
            head1.alignment = Alignment(horizontal='center', vertical='center')
            head1.value = head1.value.upper()
            head1.border = Border(top=Side(border_style='medium'),
                                  right=Side(border_style='medium'),
                                  bottom=Side(border_style='medium'),
                                  left=Side(border_style='medium'))
            worksheet.merge_cells(f'A1:{alphabet[len(inp1)-2]}1')
            worksheet.row_dimensions[1].height = 28
            worksheet.column_dimensions['G'].width = 3
            logging.debug(worksheet.max_row)
            for i in range(1, worksheet.max_row):
                worksheet[f'G{i}'].fill = PatternFill(fill_type='solid', start_color=Color(indexed=44))

    for row in ws.iter_rows(min_row=4):
        for x in range(len(inp1)):
            if (inp1[x] == 'in' or inp1[x] == 'out') and row[x].value is not None:
                data['a'][inp1[x]] = float(row[x].value)
                continue
            data['a'][inp1[x]] = row[x].value
        for y in range(8, 8 + len(inp2)):
            data['b'][inp2[y - 8]] = row[y].value

        big_cons.setdefault(data['b']['engin'], {})
        big_cons[data['b']['engin']].setdefault(data['b'][inp2[0]], {})
        big_cons[data['b']['engin']][data['b'][inp2[0]]].setdefault(data['b'][inp2[1]], {})
        cons_data = big_cons[data['b']['engin']][data['b'][inp2[0]]][data['b'][inp2[1]]]
        cons_data.setdefault('auth', data['b'][inp2[3]])
        cons_data.setdefault('out', data['b'][inp2[4]])
        cons_data.setdefault('kmA', data['b'][inp2[5]])
        cons_data.setdefault('kmB', data['b'][inp2[6]])

        try:
            ws_active = wb2[sheets[data['a'][inp1[0]].month]]
            data_a = []
            for key in data['a']:
                if key == 'in':
                    if data['a'][key] is not None:
                        supply[data['a'][inp1[0]].month] += float(data['a'][key])
                    continue
                if key == 'date':
                    data_a.append(data['a'][key].strftime('%d/%m/%Y'))
                    continue
                data_a.append(data['a'][key])

            row = ws_active.max_row
            ws_active.insert_rows(idx=ws_active.max_row)
            j = 0
            for col in ws_active.iter_cols():
                col_letter = col[0].column_letter
                if j < len(data_a):
                    ws_active[col_letter + str(row)].value = data_a[j]
                    if j == len(data_a) - 1:
                        ws_active[f'{col_letter}{row}'].number_format = '0.00'
                j += 1
            min_row = 3
            max_row = ws_active.max_row - 1
            sub_total_cell = sub_total[data['a'][inp1[0]].month]
            col_l = sub_total_cell.column_letter
            sub_total_cell.value = f'=SUM({col_l}{min_row}:{col_l}{max_row})'
        except AttributeError:
            pass

    for month in months:
        sheet = wb2[month]
        for col in sheet.iter_cols():
            try:
                col_letter = col[0].column_letter
                add = 0
                iteration = 0
                for row_number in range(1, sheet.max_row):
                    value = sheet[f'{col_letter}{row_number}'].value
                    if value is None:
                        continue
                    else:
                        add += len(str(value))
                        iteration += 1
                average = add // iteration
                sheet.column_dimensions[col_letter].width = average + 5
            except IndexError:
                continue

            for r in sheet.iter_rows(min_row=3):
                row_number = r[0].row
                sheet.row_dimensions[row_number].height = 25
                for cell in r:
                    try:
                        cell.value = cell.value.capitalize()
                    except AttributeError:
                        pass
                    cell.font = Font(color='00333333', name='Verdana', size=11)
                    cell.alignment = Alignment(vertical='center',
                                               wrap_text=True)
                    cell.border = Border(left=Side(border_style='dotted'),
                                         right=Side(border_style='dotted'),
                                         bottom=Side(border_style='dotted'))
                    if cell is not r[2]:
                        cell.alignment = Alignment(horizontal='center')
                    if cell is r[1] or cell is r[4]:
                        try:
                            cell.value = cell.value.upper()
                        except AttributeError:
                            pass
            for i in range(len(headers1)):
                cell = sheet[f'{alphabet[i]}2']
                cell.value = cell.value.upper()
                cell.font = Font(name='Verdana', bold=True, color='00111111', size=11)
                cell.border = Border(top=Side(border_style='thin', color='00333333'),
                                     right=Side(border_style='thin', color='00333333'),
                                     bottom=Side(border_style='thin', color='00333333'),
                                     left=Side(border_style='thin', color='00333333'))
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

                try:
                    if cell.value == headers1[3]:
                        sheet.column_dimensions[f'{alphabet[i]}'].width = len(cell.value)
                except TypeError:
                    pass

            sheet.row_dimensions[2].height = 26
        # Format the xlsx for better ui
        # Move cells : Identify cells to be moved
        # Move cells with their formulae
        # Make it printable
        # Auto-correct orthographic fault
        # Convert it in a pdf file for non-change

    del big_cons[None]
    write_cons_data(big_cons)

    wb2.save('rapport.xlsx')
    subprocess.Popen(['open', 'rapport.xlsx'])


monthly_rapport(gasoil_ws)
